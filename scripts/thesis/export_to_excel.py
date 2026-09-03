#!/usr/bin/env python3
"""
Runs analyze_run.py against a results directory, parses its four console tables, and
writes them into a formatted Excel workbook - so a sweep can go straight from
sim_results/ to a shareable .xlsx without anyone re-typing numbers by hand.

Does not run any simulation. Invokes analyze_run.py as a subprocess (which itself only
parses already-existing log files) and processes its stdout as text.

The four tables parsed, matched by their exact header/title text in analyze_run.py's
current print() calls - if that script's output format changes, the header markers
below need to change with it. Three of the four tables are simple whitespace-split
(every field value in those rows is a single token, verified against the f-strings that
produce them); the paired dominant-failure-reason table is the exception - its
"label (count)" cells contain an internal space, so that one table is parsed with a
regex instead of split().

Usage:
    python export_to_excel.py <results_dir> [output.xlsx]

    output.xlsx defaults to results.xlsx in the current directory.
"""
import sys
import re
import subprocess
import os
from datetime import datetime
from pathlib import Path
import configparser

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.utils import get_column_letter

HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF")
THIN = Side(style="thin", color="B0B0B0")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
PCT_FMT = "0.0%"


# ---------------------------------------------------------------------------
# Auto-increment filename to avoid overwriting existing results
# ---------------------------------------------------------------------------

def get_next_available_filename(base_name="results.xlsx"):
	"""Find next available filename: results.xlsx, results_1.xlsx, results_2.xlsx, etc."""
	if not os.path.exists(base_name):
		return base_name

	counter = 1
	while os.path.exists(f"results_{counter}.xlsx"):
		counter += 1

	return f"results_{counter}.xlsx"


# ---------------------------------------------------------------------------
# Read configuration and write CONFIG sheet
# ---------------------------------------------------------------------------

def get_config_from_file(config_path="scripts/thesis/config/default_config.properties"):
	"""Read simulation parameters from config file (handles .properties format without [DEFAULT])."""
	config = configparser.ConfigParser()
	try:
		config.read(config_path)
		if config.sections() or config.defaults():
			return config  # Successfully parsed as INI
	except configparser.MissingSectionHeaderError:
		pass  # Fall through to manual parsing

	# Manual parsing for .properties files without [DEFAULT] header
	if 'DEFAULT' not in config:
		config['DEFAULT'] = {}
	try:
		with open(config_path) as f:
			for line in f:
				line = line.strip()
				if line and not line.startswith('#'):
					key, _, value = line.partition('=')
					key = key.strip()
					value = value.strip()
					if key:
						config['DEFAULT'][key] = value
	except Exception as e:
		print(f"Warning: Could not read config file {config_path}: {e}")

	return config


def write_config_sheet(wb, config_dict):
	"""Write configuration metadata as the first sheet."""
	ws = wb.create_sheet("CONFIG", 0)  # Insert at position 0 (first sheet)

	ws["A1"] = "Simulation Configuration"
	ws["A1"].font = Font(bold=True, size=14)

	ws["A2"] = "Date & Time Run"
	ws["B2"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

	row = 4
	for key, value in config_dict.items():
		ws[f"A{row}"] = key
		ws[f"B{row}"] = value
		row += 1

	# Format: bold left column, borders
	for r in range(1, row):
		ws[f"A{r}"].font = Font(bold=True)
		for col in ["A", "B"]:
			ws[f"{col}{r}"].border = BORDER

	# Auto-size columns
	ws.column_dimensions["A"].width = 35
	ws.column_dimensions["B"].width = 50

	return ws


# ---------------------------------------------------------------------------
# Step 1: run analyze_run.py and capture its stdout
# ---------------------------------------------------------------------------

def run_analyze(results_dir: Path, script_dir: Path) -> str:
    analyze_script = script_dir / "analyze_run.py"
    apps_xml = script_dir / "config" / "applications.xml"
    if not analyze_script.exists():
        print(f"ERROR: {analyze_script} not found.")
        sys.exit(1)
    proc = subprocess.run(
        [sys.executable, str(analyze_script), str(results_dir), str(apps_xml)],
        capture_output=True, text=True, cwd=str(script_dir.parent.parent),
    )
    if proc.returncode != 0:
        print("analyze_run.py failed:")
        print(proc.stdout)
        print(proc.stderr)
        sys.exit(1)
    return proc.stdout


# ---------------------------------------------------------------------------
# Step 2: parse the four tables out of that stdout
# ---------------------------------------------------------------------------

def _pct_to_fraction(s):
    """'45.5%' -> 0.455 (fraction, for storing under a '0.0%' Excel number format).
    'n/a' -> None."""
    if s is None or s == "n/a":
        return None
    return float(s.rstrip("%")) / 100.0


def _int_or_none(s):
    if s is None or s == "n/a":
        return None
    try:
        return int(s)
    except ValueError:
        return None


def _find_block(lines, header_marker):
    """Finds the line containing header_marker, returns (data_start_index) - the
    line after the header and its dashed separator. Returns None if not found."""
    for i, ln in enumerate(lines):
        if header_marker in ln:
            # next non-empty line should be the "----" separator; data starts after it
            return i + 2
    return None


def _collect_rows(lines, start_idx):
    rows = []
    for ln in lines[start_idx:]:
        if ln.strip() == "":
            break
        rows.append(ln)
    return rows


def parse_compliance_table(stdout_text):
    """Table 1 from analyze_run.py main(): devices/policy/class/n_ok/n_fail/within%/
    MANhop%/ownFail%. Every field is a single whitespace-delimited token."""
    lines = stdout_text.splitlines()
    start = _find_block(lines, "n_ok")
    if start is None:
        return []
    rows = []
    for ln in _collect_rows(lines, start):
        parts = ln.split()
        if len(parts) != 8:
            continue
        devices, policy, cls, n_ok, n_fail, within, man_pct, own_fail = parts
        rows.append(dict(
            devices=int(devices), policy=policy, cls=cls,
            n_ok=int(n_ok), n_fail=int(n_fail),
            within=_pct_to_fraction(within), man_pct=_pct_to_fraction(man_pct),
            own_fail=_pct_to_fraction(own_fail),
        ))
    return rows


def parse_failure_breakdown_table(stdout_text):
    """Table 2: devices/policy/class/vmCap/wlanCov/mobility/bwLAN/bwMAN/bwWAN/bwGSM."""
    lines = stdout_text.splitlines()
    start = _find_block(lines, "vmCap")
    if start is None:
        return []
    rows = []
    for ln in _collect_rows(lines, start):
        parts = ln.split()
        if len(parts) != 10:
            continue
        devices, policy, cls, vmcap, wlancov, mobility, bwlan, bwman, bwwan, bwgsm = parts
        rows.append(dict(
            devices=int(devices), policy=policy, cls=cls,
            vm_capacity=int(vmcap), wlan_coverage=int(wlancov), mobility=int(mobility),
            bw_lan=int(bwlan), bw_man=int(bwman), bw_wan=int(bwwan), bw_gsm=int(bwgsm),
        ))
    return rows


def parse_paired_compliance_table(stdout_text):
    """Table 3: devices/class/CENTR within%/DECENTR within%/gap/CENTR MANhop%/
    DECENTR MANhop%. gap is dropped here - the workbook recomputes it as a live
    formula rather than trusting the Python-computed string."""
    lines = stdout_text.splitlines()
    start = _find_block(lines, "CENTR within%")
    if start is None:
        return []
    rows = []
    for ln in _collect_rows(lines, start):
        parts = ln.split()
        if len(parts) != 7:
            continue
        devices, cls, c_within, d_within, gap, c_man, d_man = parts
        rows.append(dict(
            devices=int(devices), cls=cls,
            centr_within=_pct_to_fraction(c_within), decentr_within=_pct_to_fraction(d_within),
            centr_man=_pct_to_fraction(c_man), decentr_man=_pct_to_fraction(d_man),
        ))
    return rows


_DOMINANT_RE = re.compile(r"^(\S+)\s+\((\d+)\)$")


def parse_paired_dominant_table(stdout_text):
    """Table 4: devices/class/'CENTR dominant'/'DECENTR dominant', where each dominant
    cell is text like 'mobility (50)' - two whitespace tokens for one logical value, so
    this table cannot be split() the way the other three can. Parsed with fixed-width
    slicing matching print_paired_failure_reasons()'s exact format string
    (f"{devices:>8} {cls:<12} {label:<18} {label:<18}"), then a regex splits each
    sliced cell into (reason, count)."""
    lines = stdout_text.splitlines()
    start = _find_block(lines, "CENTR dominant")
    if start is None:
        return []
    rows = []
    for ln in _collect_rows(lines, start):
        if len(ln) < 22:
            continue
        devices_s = ln[0:8].strip()
        cls = ln[9:21].strip()
        centr_cell = ln[22:40].strip()
        decentr_cell = ln[41:59].strip() if len(ln) > 41 else ""
        if not devices_s.isdigit() or not cls:
            continue
        cm = _DOMINANT_RE.match(centr_cell)
        dm = _DOMINANT_RE.match(decentr_cell)
        rows.append(dict(
            devices=int(devices_s), cls=cls,
            centr_reason=cm.group(1) if cm else centr_cell or None,
            centr_count=int(cm.group(2)) if cm else None,
            decentr_reason=dm.group(1) if dm else decentr_cell or None,
            decentr_count=int(dm.group(2)) if dm else None,
        ))
    return rows


# ---------------------------------------------------------------------------
# Step 3: write the workbook
# ---------------------------------------------------------------------------

def style_header(ws, row=1, ncols=None):
    ncols = ncols or ws.max_column
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")
    ws.freeze_panes = ws.cell(row=row + 1, column=1)


def border_data(ws, first_row=2):
    for row in ws.iter_rows(min_row=first_row):
        for cell in row:
            cell.border = BORDER


def autosize(ws):
    for col_cells in ws.columns:
        length = max((len(str(c.value)) if c.value is not None else 0) for c in col_cells)
        ws.column_dimensions[get_column_letter(col_cells[0].column)].width = max(10, length + 2)


def write_paired_comparison_sheet(wb, paired_rows):
    ws = wb.create_sheet("Paired Comparison")
    headers = ["Devices", "Service Class",
               "CENTRALIZED Compliance %", "DECENTRALIZED Compliance %", "Gap (D-C)",
               "CENTRALIZED MAN-Hop %", "DECENTRALIZED MAN-Hop %"]
    ws.append(headers)
    for r in sorted(paired_rows, key=lambda x: (x["devices"], x["cls"])):
        row_idx = ws.max_row + 1
        ws.cell(row=row_idx, column=1, value=r["devices"])
        ws.cell(row=row_idx, column=2, value=r["cls"])
        c3 = ws.cell(row=row_idx, column=3, value=r["centr_within"])
        c4 = ws.cell(row=row_idx, column=4, value=r["decentr_within"])
        for c in (c3, c4):
            c.number_format = PCT_FMT
        # Live formula, not a Python-computed value: gap = DECENTR - CENTR.
        gap_cell = ws.cell(row=row_idx, column=5,
                            value=f"=IF(OR(ISBLANK(C{row_idx}),ISBLANK(D{row_idx})),\"\",D{row_idx}-C{row_idx})")
        gap_cell.number_format = PCT_FMT
        c6 = ws.cell(row=row_idx, column=6, value=r["centr_man"])
        c7 = ws.cell(row=row_idx, column=7, value=r["decentr_man"])
        for c in (c6, c7):
            c.number_format = PCT_FMT
    style_header(ws)
    border_data(ws)
    last_row = ws.max_row
    if last_row > 1:
        # Red = small gap, green = large gap. min/max are relative to the actual data
        # range, so this adapts automatically as new sweeps are appended.
        ws.conditional_formatting.add(
            f"E2:E{last_row}",
            ColorScaleRule(start_type="min", start_color="F8696B",
                            end_type="max", end_color="63BE7B"),
        )
    autosize(ws)
    return ws, last_row


def write_failure_mechanism_sheet(wb, dominant_rows):
    ws = wb.create_sheet("Failure Mechanism")
    headers = ["Devices", "Service Class",
               "CENTRALIZED - Dominant Failure Reason", "No. of Failures",
               "DECENTRALIZED - Dominant Failure Reason", "No. of Failures"]
    ws.append(headers)
    for r in sorted(dominant_rows, key=lambda x: (x["devices"], x["cls"])):
        ws.append([r["devices"], r["cls"], r["centr_reason"], r["centr_count"],
                   r["decentr_reason"], r["decentr_count"]])
    style_header(ws)
    border_data(ws)
    autosize(ws)
    return ws


def write_compliance_detail_sheet(wb, compliance_rows):
    ws = wb.create_sheet("Compliance Detail")
    headers = ["Devices", "Policy", "Service Class", "No. Tasks Passed", "No. Tasks Failed",
               "Compliance %", "MAN-Hop %", "Failure %"]
    ws.append(headers)
    for r in sorted(compliance_rows, key=lambda x: (x["devices"], x["policy"], x["cls"])):
        row_idx = ws.max_row + 1
        ws.append([r["devices"], r["policy"], r["cls"], r["n_ok"], r["n_fail"],
                   r["within"], r["man_pct"], r["own_fail"]])
        for col in (6, 7, 8):
            ws.cell(row=row_idx, column=col).number_format = PCT_FMT
    style_header(ws)
    border_data(ws)
    autosize(ws)
    return ws


def write_failure_breakdown_sheet(wb, breakdown_rows):
    ws = wb.create_sheet("Failure Breakdown")
    headers = ["Devices", "Policy", "Service Class",
               "No. Failed (VM Overload)", "No. Failed (WLAN Loss)", "No. Failed (Device Mobility)",
               "No. Failed (LAN Bandwidth)", "No. Failed (MAN Bandwidth)",
               "No. Failed (WAN Bandwidth)", "No. Failed (GSM Bandwidth)"]
    ws.append(headers)
    for r in sorted(breakdown_rows, key=lambda x: (x["devices"], x["policy"], x["cls"])):
        ws.append([r["devices"], r["policy"], r["cls"], r["vm_capacity"], r["wlan_coverage"],
                   r["mobility"], r["bw_lan"], r["bw_man"], r["bw_wan"], r["bw_gsm"]])
    style_header(ws)
    border_data(ws)
    autosize(ws)
    return ws


def write_summary_sheet(wb, paired_last_row, results_dir):
    ws = wb.create_sheet("Summary", 0)
    ws["A1"] = "Thesis sweep summary"
    ws["A1"].font = Font(bold=True, size=16)
    ws["A2"] = f"Source: {results_dir}"
    ws["A2"].font = Font(italic=True, color="666666")

    ws["A4"] = "Largest compliance gap (DECENTRALIZED - CENTRALIZED), any device count/class:"
    ws["A4"].font = Font(bold=True)
    if paired_last_row and paired_last_row > 1:
        ws["B4"] = f"='Paired Comparison'!E2:E{paired_last_row}"
        # MAX over the same live range as above; array-safe and ignores the blank/""
        # cells the gap formula produces when a policy's data is missing for a row.
        ws["B4"] = f"=MAX('Paired Comparison'!E2:E{paired_last_row})"
    else:
        ws["B4"] = "n/a (no paired rows)"
    ws["B4"].number_format = PCT_FMT
    ws["B4"].font = Font(bold=True, size=12, color="1F4E78")

    guide_row = 6
    ws.cell(row=guide_row, column=1, value="Sheet guide").font = Font(bold=True)
    guide = [
        ("Paired Comparison", "Per-class compliance for both policies side by side, "
                               "with the gap as a live formula (=D-C) and a color scale "
                               "(red = small gap, green = large gap)."),
        ("Failure Mechanism", "The single dominant failure reason and its count, per "
                               "policy, per (device count, class) - answers 'what is "
                               "actually killing tasks' without cross-referencing two "
                               "separate tables by eye."),
        ("Compliance Detail", "Full per-policy raw numbers (n_ok, n_fail, within%, "
                               "MAN-hop%, own failure%), unpaired - for checking the "
                               "numbers a paired row was built from."),
        ("Failure Breakdown", "Full per-policy raw failure-reason counts (VM capacity, "
                               "WLAN coverage, mobility, and LAN/MAN/WAN/GSM bandwidth), "
                               "unpaired."),
    ]
    for i, (name, desc) in enumerate(guide, start=1):
        ws.cell(row=guide_row + i, column=1, value=name).font = Font(bold=True)
        ws.cell(row=guide_row + i, column=2, value=desc).alignment = Alignment(wrap_text=True)
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 90
    return ws


def build_workbook(stdout_text, results_dir, config_dict=None):
    compliance_rows = parse_compliance_table(stdout_text)
    breakdown_rows = parse_failure_breakdown_table(stdout_text)
    paired_rows = parse_paired_compliance_table(stdout_text)
    dominant_rows = parse_paired_dominant_table(stdout_text)

    warnings = []
    for label, rows in [("compliance/MAN-hop", compliance_rows),
                         ("failure-reason breakdown", breakdown_rows),
                         ("paired comparison", paired_rows),
                         ("paired dominant-reason", dominant_rows)]:
        if not rows:
            warnings.append(label)

    if warnings:
        print("WARNING: zero rows parsed for: " + ", ".join(warnings) + ".")
        print("This usually means analyze_run.py's output format changed and this")
        print("script's table-header markers / row parsers need updating to match.")
        print("Raw analyze_run.py output follows, for comparison against the parsers")
        print("in export_to_excel.py (search for _find_block calls):")
        print("-" * 70)
        print(stdout_text)
        print("-" * 70)

    if not (compliance_rows or breakdown_rows or paired_rows or dominant_rows):
        print("Nothing parsed at all - refusing to write an empty workbook.")
        sys.exit(1)

    wb = Workbook()
    wb.remove(wb.active)  # remove default blank sheet

    # Write CONFIG sheet first (at index 0)
    if config_dict:
        write_config_sheet(wb, config_dict)

    paired_ws, paired_last_row = write_paired_comparison_sheet(wb, paired_rows)
    write_failure_mechanism_sheet(wb, dominant_rows)
    write_compliance_detail_sheet(wb, compliance_rows)
    write_failure_breakdown_sheet(wb, breakdown_rows)
    # write_summary_sheet(wb, paired_last_row, results_dir)  # THESIS: disabled, not needed

    return wb


def main():
    if len(sys.argv) < 2:
        print(__doc__)
        sys.exit(1)
    results_dir = Path(sys.argv[1])

    # Use auto-increment filename if not specified
    if len(sys.argv) > 2:
        out_path = Path(sys.argv[2])
    else:
        out_path = Path(get_next_available_filename("results.xlsx"))

    script_dir = Path(__file__).resolve().parent

    # Read configuration from config file
    config_path = script_dir / "config" / "default_config.properties"
    config = get_config_from_file(str(config_path))

    # Build config dictionary to include in Excel
    config_dict = {}
    if config and 'DEFAULT' in config:
        config_dict = {
            "Simulation Time": f"{config.get('DEFAULT', 'simulation_time', fallback='N/A')} min",
            "Warm-up Period": f"{config.get('DEFAULT', 'warm_up_period', fallback='N/A')} min",
            "MAN Bandwidth": f"{config.get('DEFAULT', 'man_bandwidth', fallback='N/A')} Mbps",
            "L1 (Attractiveness)": f"{config.get('DEFAULT', 'attractiveness_L1_mean_waiting_time', fallback='N/A')} sec",
            "L2 (Attractiveness)": f"{config.get('DEFAULT', 'attractiveness_L2_mean_waiting_time', fallback='N/A')} sec",
            "L3 (Attractiveness)": f"{config.get('DEFAULT', 'attractiveness_L3_mean_waiting_time', fallback='N/A')} sec",
            "Min Device Count": config.get('DEFAULT', 'min_number_of_mobile_devices', fallback='N/A'),
            "Max Device Count": config.get('DEFAULT', 'max_number_of_mobile_devices', fallback='N/A'),
            "Device Count Step": config.get('DEFAULT', 'mobile_device_counter_size', fallback='N/A'),
            "VM MIPS Capacity": f"{config.get('DEFAULT', 'mips_for_cloud_vm', fallback='N/A')} MIPS",
            "Region Count": config.get('DEFAULT', 'region_count', fallback='N/A'),
            "Centralized Decision Delay": f"{config.get('DEFAULT', 'centralized_decision_delay', fallback='N/A')} sec",
            "Decentralized Decision Delay": f"{config.get('DEFAULT', 'decentralized_decision_delay', fallback='N/A')} sec",
            "Orchestrator Policies": config.get('DEFAULT', 'orchestrator_policies', fallback='N/A'),
            "Simulation Scenarios": config.get('DEFAULT', 'simulation_scenarios', fallback='N/A'),
        }

    stdout_text = run_analyze(results_dir, script_dir)
    wb = build_workbook(stdout_text, results_dir, config_dict)
    wb.save(out_path)
    print(f"Wrote {out_path.resolve()}")


if __name__ == "__main__":
    main()
